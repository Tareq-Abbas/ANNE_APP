from openpyxl.workbook import Workbook
from pyzbar import pyzbar
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.graphics.texture import Texture
from kivy.clock import Clock
import cv2
from kivy.core.window import Window
from kivy.properties import ObjectProperty, ListProperty , NumericProperty, StringProperty
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#from android.permissions import request_permissions, Permission
import numpy as np


Window.clearcolor = (1, 1, 1, 1) #when we create an app the first time by adding the command run() we get black window here we will get white one
Window.size = (500, 660)



class WindowManager(ScreenManager):
    pass


class StartWindow(Screen):
    pass


class CamWindow(Screen):

    def cam_add_to_history(self, val):
        filename = "history.xlsx"
        is_added = False
        #new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        for row in range(1, ws.max_row + 1):
            if val != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_added = True

        if not is_added:
            ws.append([val])

        wb.save(filename)

    def __init__(self, **kwargs):
        #this class inherit from Screen, the first step in __init__ for this class is we call the __init__ from Screen
        # and we get all the information from it
        super().__init__(**kwargs)
        self.capture = cv2.VideoCapture(0)  # changed to 0 for my laptop

        self.schedule = None
        self.barcode_info = ""

    def on_enter(self, *args):
        self.schedule = Clock.schedule_interval(self.load_video, 1.0 / 30.0)

    def load_video(self, *args):
        ret, frame = self.capture.read()
        # Frame initialize
        # self.image_frame = frame
        #buffer = cv2.flip(self.read_barcodes(frame), 0).tobytes()
        buffer = np.flip(self.read_barcodes(frame), 0).tobytes()
        texture = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt='bgr')
        texture.blit_buffer(buffer, colorfmt='bgr', bufferfmt='ubyte')
        self.ids.video.texture = texture  # id video is defined in kv




    def read_barcodes(self,frame):
        barcodes = pyzbar.decode(frame)
        for barcode in barcodes:
            x, y, w, h = barcode.rect
            # 1
            self.barcode_info = barcode.data.decode('utf-8')
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            print(self.barcode_info)

            # 2
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, self.barcode_info, (x + 6, y - 6), font, 2.0, (255, 255, 255), 1)
            # 3
            with open("barcode_result.txt", mode='w') as file:
                file.write("Recognized Barcode:" + self.barcode_info)
                # return the bounding box of the barcode
        return frame




    #def stop(self):
        #if self.schedule:
            #self.schedule.cancel()
            #self.schedule = None

    #def on_leave(self, *args):
        #self.stop()




    def get_value_from_book(self):
        self.inf = load_workbook('Book4.xlsx')
        self.inf = self.inf.active
        self.li = []
        self.count = 0
        self.founded_cer= ""

        ref_to_data_screen = self.manager.get_screen("data")
        #ref_to_forth_screen = self.manager.get_screen("history")
        # self.ids.score.text = ref_to_other_screen.ids.entered_value.text

        for row in range(1, self.inf.max_row + 1):
            if self.barcode_info == self.inf[get_column_letter(1) + str(row)].value:
                ref_to_data_screen.ids.product_name.text = self.inf[get_column_letter(2) + str(row)].value
                self.cam_add_to_history(self.inf[get_column_letter(2) + str(row)].value)
                for col in range(3, 7):
                    #self.li.append(inf[get_column_letter(col) + str(row)].value)
                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.count +=1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)
                        break


                for col in range(7, self.inf.max_column + 1):
                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.count += 1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)


        if self.count > 0:
            ref_to_data_screen.ids.score.text = str(self.count)


            # print(float(self.ids.score.text ))
            #self.ids.product_name.text = ref_to_forth_screen.ids.forth_product.text = str(self.li[0])
            #self.ids.producer.text = ref_to_forth_screen.ids.forth_producer.text = str(self.li[1])
            print(self.li)
            for cer in self.li:
                self.founded_cer += cer + "\n"

            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
                  # this should be multiplied by 60 (we have max is 6 and we have the circle 360 degree )
            if (self.count * (360 /16)) <= 120:
                ref_to_data_screen.bar_color = [1, 0, 0, 1]
            elif 120 <= (self.count * (360 /16)) < 240:
                ref_to_data_screen.bar_color = [0, 0, 1, 1]
            else:
                ref_to_data_screen.bar_color = [0, 1, 0, 1]


        else:
            ref_to_data_screen.ids.score.text = str(self.count)

        # now get info for history
        self.li_history = []
        self.inf_history = load_workbook('history.xlsx')
        self.inf_history = self.inf_history.active
        self.founded_cer_history = ""
        print(self.inf_history.max_row )
        for row in range(2, self.inf_history.max_row + 1):
            self.li_history.append(self.inf_history[get_column_letter(1) + str(row)].value)

        if not self.li_history:
            for cer in self.li_history:
                self.founded_cer_history += cer + "\n"

        self.manager.get_screen("history").ids.old_history.text = self.founded_cer_history




class TypeScan(Screen):

    def add_to_history(self, val):
        filename = "history.xlsx"
        is_added = False

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        for row in range(1, ws.max_row + 1):
            if val != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_added = True

        if not is_added:
            ws.append([val])

        wb.save(filename)


    def get_value_from_book(self):
        self.inf = load_workbook('Book4.xlsx')
        self.inf = self.inf.active
        self.li = []
        self.count = 0
        self.founded_cer = ""

        ref_to_data_screen = self.manager.get_screen("data")
        #ref_to_forth_screen = self.manager.get_screen("history")
        # self.ids.score.text = ref_to_other_screen.ids.entered_value.text

        for row in range(1, self.inf.max_row + 1):
            if self.ids.entered_value.text == self.inf[get_column_letter(1) + str(row)].value:
                ref_to_data_screen.ids.product_name.text = self.inf[get_column_letter(2) + str(row)].value
                self.add_to_history(self.inf[get_column_letter(2) + str(row)].value)
                for col in range(3, 7):

                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.count +=1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)
                        break


                for col in range(7, self.inf.max_column + 1):
                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.count += 1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)


        if self.count > 0:
            ref_to_data_screen.ids.score.text = str(self.count)

            # print(float(self.ids.score.text ))
            #self.ids.product_name.text = ref_to_forth_screen.ids.forth_product.text = str(self.li[0])
            #self.ids.producer.text = ref_to_forth_screen.ids.forth_producer.text = str(self.li[1])
            print(self.li)
            for cer in self.li:
                self.founded_cer += cer + "\n"

            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
                  # this should be multiplied by 60 (we have max is 6 and we have the circle 360 degree )
            if (self.count * (360 /16)) <= 120:
                ref_to_data_screen.bar_color = [1, 0, 0, 1]
            elif 120 <= (self.count * (360 /16)) < 240:
                ref_to_data_screen.bar_color = [0, 0, 1, 1]
            else:
                ref_to_data_screen.bar_color = [0, 1, 0, 1]


        else:
            ref_to_data_screen.ids.score.text = str(self.count)


        # now send info for history
        self.li_history = []
        self.inf_history = load_workbook('history.xlsx')
        self.inf_history = self.inf_history.active
        self.founded_cer_history = ""

        for row in range(2, self.inf_history.max_row + 1):
            self.li_history.append(self.inf_history[get_column_letter(1) + str(row)].value)
        #print(self.li_history)
        if self.li_history:
            for cer in self.li_history:
                self.founded_cer_history += cer + "\n"
        ref_to_history_screen = self.manager.get_screen("history")
        ref_to_history_screen.ids.old_history.text = self.founded_cer_history
    #pass

class History(Screen):
    old_history = ObjectProperty(None)


    #pass

class Favourites(Screen):
    my_favourites =ObjectProperty(None)
    #pass

class DataWindow(Screen):
    product_name = ObjectProperty(None)
    score = ObjectProperty(None)
    bar_color = ListProperty([0.6, 0.6, 0.6, 1])  # we initialize it here
    sc_lab = ObjectProperty(None)

    def add_to_favourites(self):
        filename = "favourites.xlsx"
        is_added = False
        #new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        for row in range(1, ws.max_row + 1):
            if self.ids.product_name.text != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_added = True

        if not is_added:
            ws.append([self.ids.product_name.text])

        self.li_favourites = []
        self.founded_cer_favourites = ""

        for row in range(2, ws.max_row + 1):
            self.li_favourites.append(ws[get_column_letter(1) + str(row)].value)
        # print(self.li_history)
        if self.li_favourites:
            for cer in self.li_favourites:
                self.founded_cer_favourites += cer + "\n"
        ref_to_favourites_screen = self.manager.get_screen("favourites")
        ref_to_favourites_screen.ids.my_favourites.text = self.founded_cer_favourites

        wb.save(filename)



    def remove_from_favourites(self):
        filename = "favourites.xlsx"
        is_deleted = 0
        # new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        for row in range(1, ws.max_row + 1):
            if self.ids.product_name.text != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_deleted = row

        if is_deleted != 0:
            ws.delete_rows(is_deleted)


        self.li_favourites = []
        self.founded_cer_favourites = ""

        for row in range(2, ws.max_row + 1):
            self.li_favourites.append(ws[get_column_letter(1) + str(row)].value)
        # print(self.li_history)
        if self.li_favourites:
            for cer in self.li_favourites:
                self.founded_cer_favourites += cer + "\n"
        ref_to_favourites_screen = self.manager.get_screen("favourites")
        ref_to_favourites_screen.ids.my_favourites.text = self.founded_cer_favourites

        wb.save(filename)

    #pass


returned_value= Builder.load_file('anne1.kv')


class ANNE(App):
    def build(self):
        self.title = 'ANNE'
        #request_permissions([
            #Permission.CAMERA,
            #Permission.WRITE_EXTERNAL_STORAGE,
            #Permission.READ_EXTERNAL_STORAGE
        #])
        return returned_value



if __name__ == '__main__':
    ANNE().run()
