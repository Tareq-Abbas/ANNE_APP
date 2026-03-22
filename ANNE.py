from kivy.clock import mainthread, Clock
from kivy.properties import ObjectProperty, ListProperty #, NumericProperty, StringProperty
from kivy.lang import Builder
from kivymd.app import MDApp
from kivy.uix.screenmanager import ScreenManager, Screen
from camera4kivy import Preview
from PIL import Image
from pyzbar.pyzbar import decode
#from kivy.utils import platform
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



certifications_links = ['[ref=https://rspo.org/as-an-organisation/certification/supply-chains/]',
                        '[ref=https://controlunion-germany.com/de/certification-programs/eu-okolandbau-eu-bio-siegel]',
                        '[ref=https://www.bmel.de/DE/themen/landwirtschaft/oekologischer-landbau/bio-siegel.html]',
                        '[ref=https://demeter.net/certification/]',
                        '[ref=https://www.bioland.de/verbraucher]',
                        '[ref=https://www.msc.org/de]',
                        '[ref=https://proveg.com/de/ueber-uns/]',
                        '[ref=https://www.avocadostore.de/wissenswert/siegel/veganblume]',
                        '[ref=https://www.diqp.eu/guetesiegel-erkennen/]',
                        '[ref=https://www.rainforest-alliance.org/insights/what-does-rainforest-alliance-certified-mean/]',
                        '[ref=https://www.rainforest-alliance.org/de/utz/]',
                        '[ref=https://www.blauer-engel.de/de]',
                        '[ref=https://www.budni.de/unternehmen/article/mikroplastik]',
                        '[ref=https://newsroom.kunststoffverpackungen.de/2021/07/05/was-ist-pir-und-was-pcr/]',
                        '[ref=https://www.wwf.de/themen-projekte/waelder/verantwortungsvollere-waldnutzung/fsc-was-ist-das]',
                        '[ref=https://de.wikipedia.org/wiki/Nachf%C3%BCllpackung]']



class WindowManager(ScreenManager):
    # this function is added to deal with the icons in navigation bar
    # usually to change the screen we initialize "current" with the "name of the wanted screen"
    # but the icons in nav_bar accepts only functions so we this function to move to "screen"
    def change_screen(self, screen):
        # the same as in .kv: app.root.current = screen
        self.current = screen


class StartWindow(Screen):
    pass


class CamWindow(Screen):

    def get_value_from_book(self):
        self.inf = load_workbook('Book4.xlsx')
        self.inf = self.inf.active
        #now we create a list to store all founded certificates
        self.li = []
        #number of founded certificates
        self.count = 0
        self.founded_cer = ""

        ref_to_data_screen = self.manager.get_screen("data")
        # ref_to_forth_screen = self.manager.get_screen("history")
        # self.ids.score.text = ref_to_other_screen.ids.entered_value.text

        for row in range(1, self.inf.max_row + 1):
            if self.ids.ti.text == self.inf[get_column_letter(1) + str(row)].value:
                ref_to_data_screen.ids.product_name.text = self.inf[get_column_letter(2) + str(row)].value
                # now if the product in our product database we want to store it in history.xlsx if not existed
                # it means the customer is scanning products from other supermarkets so no need to store that in history
                self.cam_add_to_history(self.inf[get_column_letter(2) + str(row)].value) # add to history when founded
                for col in range(3, 7):
                    # self.li.append(inf[get_column_letter(col) + str(row)].value)
                    if not self.inf[get_column_letter(col) + str(row)].value == None: # we need to see which certificates are marked to add
                        self.founded_cer +=  certifications_links[0] + str(self.count + 1) + "- " + self.inf[get_column_letter(col) + str(2)].value + '[/ref]' + "\n" + '\n'
                        self.count += 1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)
                        break

                for col in range(7, self.inf.max_column + 1):
                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.founded_cer += certifications_links[col - 6] + str(
                            self.count + 1) + "- " + self.inf[
                                                get_column_letter(col) + str(2)].value + '[/ref]' + "\n" + '\n'
                        self.count += 1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)

        if self.count > 0:
            ref_to_data_screen.ids.score.text = "count: " + str(self.count)

            # print(float(self.ids.score.text ))
            # self.ids.product_name.text = ref_to_forth_screen.ids.forth_product.text = str(self.li[0])
            # self.ids.producer.text = ref_to_forth_screen.ids.forth_producer.text = str(self.li[1])
            #print(self.li)
            #self.ind = 0
            #for cer in self.li:
                #self.ind += 1
                #self.founded_cer += str(self.ind) + "- " + cer + "\n"


            # the progress bar is 360 degree
            # and we have maximum 16 certificates
            # so the fomula is count * (360 / 16)
            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
            # this should be multiplied by 60 (we have max is 6 and we have the circle 360 degree )
            if (self.count * (360 / 16)) <= 120:
                ref_to_data_screen.bar_color = [1, 0, 0, 1]
            elif 120 <= (self.count * (360 / 16)) < 240:
                ref_to_data_screen.bar_color = [0, 0, 1, 1]
            else:
                ref_to_data_screen.bar_color = [0, 1, 0, 1]

            self.ids.ti.text = ""

        else:
            ref_to_data_screen.ids.score.text = str(self.count)
            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
            self.ids.ti.text = ""

    # we want to add the searched product to the history.xlsx
    def cam_add_to_history(self, val):
        filename = "history.xlsx"
        is_added = False
        # new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        for row in range(1, ws.max_row + 1):
            if val != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_added = True

        if not is_added:
            ws.append([val])

        wb.save(filename)


    def on_kv_post(self, obj):
        try:
            self.ids.preview.connect_camera( enable_photo = False, enable_analyze_pixels=True, default_zoom=0.0)
        except Exception as e:
            print(e)
            # Error: 'super' object has no attribute '__getattr__'

        else:
            self.ids.preview.connect_camera( enable_photo = False, enable_analyze_pixels=True, default_zoom=0.0)

            # Error: AttributeError: 'ScanScreen' object has no attribute 'root'

    @mainthread
    def got_result(self, result):
        self.ids.ti.text = str(result)



class ScanAnalyze(Preview):
    extracted_data = ObjectProperty(None)

    def analyze_pixels_callback(self, pixels, image_size, image_pos, scale, mirror):
        #use it to read image data from a binary file or memory buffer and create a PIL image so that we can decode it
        pimage = Image.frombytes(mode='RGBA', size=image_size, data=pixels)
        #decode function takes the image object, and decodes multiple barcodes or QR Codes in that image

        list_of_all_barcodes = decode(pimage)
        barcode_info = ""
        ## decode() function returns a list information (data, type, rect, and polygon), we need
        # only the data, and to decode it using utf8 to get a string.
        # type :Only useful for barcodes as it outlines the barcode format.
        #  Rect: object which represents the captured localization area
        # polygon : A list of Point instances which represents the barcode or QR Code.
        for barcode in list_of_all_barcodes:
            barcode_info = barcode.data.decode('utf-8')

        if list_of_all_barcodes:
            if self.extracted_data:
                self.extracted_data(barcode_info)
            else:
                print("Not found")



class TypeScan(Screen):
    def add_to_history(self, val):
        filename = "history.xlsx"
        is_added = False

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        for row in range(1, ws.max_row + 1):
            if val != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_added = True

        if not is_added:
            ws.append([val])

        wb.save(filename)

    def get_value_from_book(self):
        self.inf = load_workbook('Book4.xlsx')
        self.inf = self.inf.active
        self.li = []
        self.count = 0
        self.founded_cer = ""

        ref_to_data_screen = self.manager.get_screen("data")
        # ref_to_forth_screen = self.manager.get_screen("history")
        # self.ids.score.text = ref_to_other_screen.ids.entered_value.text

        for row in range(1, self.inf.max_row + 1):
            if self.ids.entered_value.text == self.inf[get_column_letter(1) + str(row)].value:
                ref_to_data_screen.ids.product_name.text = self.inf[get_column_letter(2) + str(row)].value
                self.add_to_history(self.inf[get_column_letter(2) + str(row)].value)
                for col in range(3, 7):

                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.founded_cer +=  certifications_links[0] + str(self.count + 1) + "- " + self.inf[get_column_letter(col) + str(2)].value + '[/ref]' + "\n" + '\n'
                        self.count += 1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)
                        break

                for col in range(7, self.inf.max_column + 1):
                    if not self.inf[get_column_letter(col) + str(row)].value == None:
                        self.founded_cer +=  certifications_links[col - 6] + str(self.count + 1) + "- " + self.inf[get_column_letter(col) + str(2)].value + '[/ref]' + "\n" + '\n'
                        self.count += 1
                        self.li.append(self.inf[get_column_letter(col) + str(2)].value)

        if self.count > 0:
            ref_to_data_screen.ids.score.text = str(self.count)

            # print(float(self.ids.score.text ))
            # self.ids.product_name.text = ref_to_forth_screen.ids.forth_product.text = str(self.li[0])
            # self.ids.producer.text = ref_to_forth_screen.ids.forth_producer.text = str(self.li[1])
            #print(self.li)
            #self.ind = 0
            #for cer in self.li:
                #self.ind += 1
                #self.founded_cer += str(self.ind) + "- " + cer + "\n"

            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
            # this should be multiplied by 60 (we have max is 6 and we have the circle 360 degree )
            if (self.count * (360 / 16)) <= 120:
                ref_to_data_screen.bar_color = [1, 0, 0, 1]
            elif 120 <= (self.count * (360 / 16)) < 240:
                ref_to_data_screen.bar_color = [0, 0, 1, 1]
            else:
                ref_to_data_screen.bar_color = [0, 1, 0, 1]

            #self.count = 0
            self.ids.entered_value.text = ""


        else:
            ref_to_data_screen.ids.score.text = str(self.count)
            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
            self.ids.entered_value.text = ""


class Favourites(Screen):
    my_favourites = ObjectProperty(None)

    # when clicking refresh icon we want to get all the stored data from history.xlsx and show on screen
    # using viewscroll
    def getData(self):
        filename = "favourites.xlsx"
        # is_added = False
        # new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        self.li_favourites = []
        self.founded_cer_favourites = ""


        # we start from the second row from the data base because the first one is the title
        for row in range(2, ws.max_row + 1):
            self.li_favourites.append(ws[get_column_letter(1) + str(row)].value)
        # print(self.li_history)


        # if the data base not empty we want all the products name from it
        if self.li_favourites:
            for cer in self.li_favourites:
                self.founded_cer_favourites += "> " + cer + "\n"
        ref_to_favourites_screen = self.manager.get_screen("favourites")
        ref_to_favourites_screen.ids.my_favourites.text = self.founded_cer_favourites


        wb.save(filename)
        #pass


class History(Screen):
    old_history = ObjectProperty(None)
    # when clicking refresh icon we want to get all the stored data from history.xlsx and show on screen
    # using viewscroll
    def getData(self):
        filename = "history.xlsx"
        #is_added = False
        # new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        self.li_history = []
        self.founded_cer_history = ""


        # we start from the second row from the data base because the first one is the title
        for row in range(2, ws.max_row + 1):
            self.li_history.append(ws[get_column_letter(1) + str(row)].value)
        # print(self.li_history)

        # if the database not empty then we want to get all the products from it
        if self.li_history:
            for cer in self.li_history:
                self.founded_cer_history += "> " + cer + "\n"
        ref_to_history_screen = self.manager.get_screen("history")
        ref_to_history_screen.ids.old_history.text = self.founded_cer_history

        wb.save(filename)
        #pass


class DataWindow(Screen):
    product_name = ObjectProperty(None)
    score = ObjectProperty(None)
    bar_color = ListProperty([0.6, 0.6, 0.6, 1])  # we initialize it here
    sc_lab = ObjectProperty(None)
    def add_to_favourites(self):
        filename = "favourites.xlsx"
        is_added = False
        # new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)

        # see if product name we get exists in the favourites xlsx
        # if existed store then tell me it is true
        for row in range(1, ws.max_row + 1):
            if self.ids.product_name.text != ws[get_column_letter(1) + str(row)].value: # we want to add the founded product to favourites if it is not saved before
                continue
            else:
                is_added = True


        # if it in not added (is_added false) then we want to add it
        if not is_added:
            ws.append([self.ids.product_name.text])


        # save all changes in data base before ending the function
        wb.save(filename)

    def remove_from_favourites(self):
        filename = "favourites.xlsx"
        is_deleted = 0

        # new_row = ['1', '2', '3']

        # Confirm file exists.
        # If not, create it, add headers, then append new data
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]  # select first worksheet
        except FileNotFoundError:
            headers_row = ['Products']
            wb = Workbook()
            ws = wb.active
            ws.append(headers_row)


        # see if product name we get exists in the favourites xlsx
        # if existed store the row
        for row in range(1, ws.max_row + 1):
            if self.ids.product_name.text != ws[get_column_letter(1) + str(row)].value:
                continue
            else:
                is_deleted = row
        # if the row not 0 then we found the product name in the data base so we want to delete that specific row
        if is_deleted != 0:
            ws.delete_rows(is_deleted)


        #save all changes before ending the function
        wb.save(filename)
    #pass


class Anne(MDApp):
    # the next function gives information about the app like theme_style and color
    # and here we add the permissions needed in the android device to deal with camera and external storage (xlsx files)
    def build(self):
        self.theme_cls.theme_style = "Light"
        self.theme_cls.ptimary_palette = "BlueGray"
        #if platform == 'android':
            #from android.permissions import request_permissions, Permission
            #request_permissions([Permission.WRITE_EXTERNAL_STORAGE, Permission.CAMERA, Permission.RECORD_AUDIO])
        return Builder.load_file('anne1.kv')

    # to connect the camera to the second screen we use the next 2 functions
    # (look at the link on the 2nd line in this file)
    def on_start(self):
        Clock.schedule_once(self.connect_camera)

    def connect_camera(self, dt):
        secoundWindow = self.root.get_screen('cam')
        secoundWindow.ids.preview.connect_camera( enable_photo = False, enable_analyze_pixels=True, default_zoom=0.0)

# running the app
if __name__ == '__main__':
    Anne().run()