
#kivy camera application with opencv in android shows black screen
# if it does not work use camera4kivy https://pypi.org/project/camera4kivy/
#it works on all os, when i click on button info , i take a picture then i use pyzbar to read the barcode



from kivy.uix.camera import Camera
from kivy.uix.boxlayout import BoxLayout
import numpy as np
from pyzbar import pyzbar
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.graphics.texture import Texture
from kivy.clock import Clock
import cv2
from kivy.core.window import Window
from kivy.properties import ObjectProperty, ListProperty, NumericProperty, StringProperty
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

Window.clearcolor = (1,1,1,1) #when we create an app the first time by adding the command run() we get black window here we will get white one
Window.size = (500, 660)
inf = load_workbook('Book4.xlsx')
inf = inf.active

returned_value = Builder.load_file('myapplayout.kv')


class WindowManager(ScreenManager):
    pass


class StartWindow(Screen):
    pass


class CamWindow(Screen):
    camera_resolution = (640, 480)
    cam_ratio = camera_resolution[0] / camera_resolution[1]
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        #self.capture = cv2.VideoCapture(0)  # changed to 0 for my laptop
        #self.schedule = None
        #self.barcode_info= ""
        self.counter = 0



    def on_start(self):
        Clock.schedule_once(self.get_frame, 5)



    def read_barcodes(self,frame):
        barcodes = pyzbar.decode(frame)
        for barcode in barcodes:
            x, y, w, h = barcode.rect
            # 1
            self.barcode_info = barcode.data.decode('utf-8')
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            print(self.barcode_info)

            # 2
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, self.barcode_info, (x + 6, y - 6), font, 2.0, (255, 255, 255), 1)
            # 3
            with open("barcode_result.txt", mode='w') as file:
                file.write("Recognized Barcode:" + self.barcode_info)
                # return the bounding box of the barcode
        return frame

    def get_frame(self, dt):
        cam = self.root.ids.a_cam
        image_object = cam.export_as_image(scale=round((400 / int(cam.height)), 2))
        w, h = image_object._texture.size
        frame = self.read_barcodes(np.frombuffer(image_object._texture.pixels, 'uint8').reshape(h, w, 4))
        gray = cv2.cvtColor(frame, cv2.COLOR_RGBA2GRAY)
        self.root.ids.frame_counter.text = f'frame: {self.counter}'
        self.counter += 1
        Clock.schedule_once(self.get_frame, 0.25)


    def read_barcodes(self,frame):
        barcodes = pyzbar.decode(frame)
        for barcode in barcodes:
            x, y, w, h = barcode.rect
            # 1
            self.barcode_info = barcode.data.decode('utf-8')
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            print(self.barcode_info)

            # 2
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, self.barcode_info, (x + 6, y - 6), font, 2.0, (255, 255, 255), 1)
            # 3
            with open("barcode_result.txt", mode='w') as file:
                file.write("Recognized Barcode:" + self.barcode_info)
                # return the bounding box of the barcode
        return frame


    def get_value_from_book(self):
        self.li = []
        self.count = 0
        self.founded_cer= ""

        ref_to_data_screen = self.manager.get_screen("data")
        #ref_to_forth_screen = self.manager.get_screen("history")
        # self.ids.score.text = ref_to_other_screen.ids.entered_value.text

        for row in range(1, inf.max_row + 1):
            if self.barcode_info == inf[get_column_letter(1) + str(row)].value:
                ref_to_data_screen.ids.product_name.text = inf[get_column_letter(2) + str(row)].value
                for col in range(3, 7):
                    #self.li.append(inf[get_column_letter(col) + str(row)].value)
                    if not inf[get_column_letter(col) + str(row)].value == None:
                        self.count +=1
                        self.li.append(inf[get_column_letter(col) + str(2)].value)
                        break


                for col in range(7, inf.max_column + 1):
                    if not inf[get_column_letter(col) + str(row)].value == None:
                        self.count += 1
                        self.li.append(inf[get_column_letter(col) + str(2)].value)


        if self.count > 0:
            ref_to_data_screen.ids.score.text = str(self.count)
            # print(float(self.ids.score.text ))
            #self.ids.product_name.text = ref_to_forth_screen.ids.forth_product.text = str(self.li[0])
            #self.ids.producer.text = ref_to_forth_screen.ids.forth_producer.text = str(self.li[1])
            print(self.li)
            for cer in self.li:
                self.founded_cer += cer + "\n"

            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
                  # this should be multiplied by 60 (we have max is 6 and we have the circle 360 degree )
            if (self.count * (360 /16)) <= 120:
                ref_to_data_screen.bar_color = [1, 0, 0, 1]
            elif 120 <= (self.count * (360 /16)) < 240:
                ref_to_data_screen.bar_color = [0, 0, 1, 1]
            else:
                ref_to_data_screen.bar_color = [0, 1, 0, 1]


        else:
            ref_to_data_screen.ids.score.text = str(self.count)




class TypeScan(Screen):


    def get_value_from_book(self):
        self.li = []
        self.count = 0
        self.founded_cer = ""

        ref_to_data_screen = self.manager.get_screen("data")
        #ref_to_forth_screen = self.manager.get_screen("history")
        # self.ids.score.text = ref_to_other_screen.ids.entered_value.text

        for row in range(1, inf.max_row + 1):
            if self.ids.entered_value.text == inf[get_column_letter(1) + str(row)].value:
                ref_to_data_screen.ids.product_name.text = inf[get_column_letter(2) + str(row)].value
                for col in range(3, 7):

                    if not inf[get_column_letter(col) + str(row)].value == None:
                        self.count +=1
                        self.li.append(inf[get_column_letter(col) + str(2)].value)
                        break


                for col in range(7, inf.max_column + 1):
                    if not inf[get_column_letter(col) + str(row)].value == None:
                        self.count += 1
                        self.li.append(inf[get_column_letter(col) + str(2)].value)


        if self.count > 0:
            ref_to_data_screen.ids.score.text = str(self.count)
            # print(float(self.ids.score.text ))
            #self.ids.product_name.text = ref_to_forth_screen.ids.forth_product.text = str(self.li[0])
            #self.ids.producer.text = ref_to_forth_screen.ids.forth_producer.text = str(self.li[1])
            print(self.li)
            for cer in self.li:
                self.founded_cer += cer + "\n"

            ref_to_data_screen.ids.score_progress_bar.progress = self.count * (360 / 16)
            ref_to_data_screen.ids.sc_lab.text = self.founded_cer
                  # this should be multiplied by 60 (we have max is 6 and we have the circle 360 degree )
            if (self.count * (360 /16)) <= 120:
                ref_to_data_screen.bar_color = [1, 0, 0, 1]
            elif 120 <= (self.count * (360 /16)) < 240:
                ref_to_data_screen.bar_color = [0, 0, 1, 1]
            else:
                ref_to_data_screen.bar_color = [0, 1, 0, 1]


        else:
            ref_to_data_screen.ids.score.text = str(self.count)
    #pass

class History(Screen):
    pass

class Favourites(Screen):
    pass

class DataWindow(Screen):
    product_name = ObjectProperty(None)
    score = ObjectProperty(None)
    bar_color = ListProperty([0.6, 0.6, 0.6, 1])  # we initialize it here
    sc_lab =  ObjectProperty(None) #StringProperty('')



    #pass




class MyApplayout(App):
    def build(self):
        self.title = 'ANNE'
        return returned_value


if __name__ == "__main__":
    MyApplayout().run()





















'''
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.graphics.texture import Texture
from kivy.uix.camera import Camera
from kivy.lang import Builder
import numpy as np
import cv2

Builder.load_file("myapplayout.kv")

class AndroidCamera(Camera):
    camera_resolution = (640, 480)
    counter = 0

    def _camera_loaded(self, *largs):
        self.texture = Texture.create(size=np.flip(self.camera_resolution), colorfmt='rgb')
        self.texture_size = list(self.texture.size)

    def on_tex(self, *l):
        if self._camera._buffer is None:
            return None
        frame = self.frame_from_buf()
        self.frame_to_screen(frame)
        super(AndroidCamera, self).on_tex(*l)

    def frame_from_buf(self):
        w, h = self.resolution
        frame = np.frombuffer(self._camera._buffer.tostring(), 'uint8').reshape((h + h // 2, w))
        frame_bgr = cv2.cvtColor(frame, 93)
        return np.rot90(frame_bgr, 3)

    def frame_to_screen(self, frame):
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        cv2.putText(frame_rgb, str(self.counter), (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_AA)
        self.counter += 1
        flipped = np.flip(frame_rgb, 0)
        buf = flipped.tostring()
        self.texture.blit_buffer(buf, colorfmt='rgb', bufferfmt='ubyte')

class MyLayout(BoxLayout):
    pass

class MyApp(App):
    def build(self):
        return MyLayout()

if __name__ == '__main__':
    MyApp().run() 
    
    
'''